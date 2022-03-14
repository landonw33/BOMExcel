import openpyxl

bom_file = openpyxl.load_workbook("bom.xlsx")
product_list = bom_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_ten_inv = {}
life_cycle_status = {}
message = {'CHECK SOON'}

# print(product_list.max_row)
for product_row in range(2, product_list.max_row + 1):
    manufacturer_name = product_list.cell(product_row, 1).value
    part_number = product_list.cell(product_row, 2).value
    inventory = product_list.cell(product_row, 3).value
    price = product_list.cell(product_row, 4).value
    life_cycle = product_list.cell(product_row, 5).value
    bom_price = product_list.cell(product_row, 6)
    needs_attention = product_list.cell(product_row, 7)

    # calculation number of products per supplier
    if manufacturer_name in products_per_supplier:
        current_num_products = products_per_supplier[manufacturer_name]
        products_per_supplier[manufacturer_name] = current_num_products + 1
    else:
        products_per_supplier[manufacturer_name] = 1

    # find total inv per supplier
    if manufacturer_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(manufacturer_name)
        total_value_per_supplier[manufacturer_name] = current_total_value + (inventory * price)
    else:
        total_value_per_supplier[manufacturer_name] = inventory * price

    # what products are over 3 on lifecycle
    if life_cycle > 3:
        life_cycle_status[manufacturer_name] = int(life_cycle)
        # message = f"{item}"

    # Adds total cost to a new xlsx sheet
    bom_price.value = inventory*price
    if life_cycle > 3:
        needs_attention.value = 'CHECK SOON'

for item in products_per_supplier:
    print(f"Manufacturer{item} has {products_per_supplier[item]} items on this BOM.")
    print(f"Manufacturer{item}'s total cost is {total_value_per_supplier[item]}")
    print("\n")

print(f"Items to be checked")
for manufacturer_name in life_cycle_status:
    print(f"Manufacturer  {manufacturer_name} has a part with a lifecycle of {life_cycle_status[manufacturer_name]}")

# saving to our excel sheet
bom_file.save("bom_new.xlsx")
